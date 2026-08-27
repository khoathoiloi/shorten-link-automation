# -*- coding: utf-8 -*-
import os
import sys
import re
import time
import random
import urllib.request
import json
import subprocess
import openpyxl

CURRENT_VERSION = "1.0.3"
GITHUB_REPO = "khoathoiloi/shorten-link-automation"

# Đảm bảo hiển thị Tiếng Việt trên Windows console
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stdin.reconfigure(encoding="utf-8")
    except Exception:
        pass

from playwright.sync_api import sync_playwright

TARGET_URL = "https://shorten-link-swart.vercel.app/"

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
    CURRENT_EXE = sys.executable
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    CURRENT_EXE = None

USER_DATA_DIR = os.path.join(APP_DIR, "user_data")

def is_newer_version(latest, current):
    try:
        l_parts = [int(p) for p in latest.split('.')]
        c_parts = [int(p) for p in current.split('.')]
        return l_parts > c_parts
    except Exception:
        return latest != current

def perform_update(download_url, new_version):
    """Tải file exe mới và tự động thay thế file cũ"""
    if not CURRENT_EXE:
        print(f"⚠️ Bạn đang chạy mã nguồn Python. Vui lòng tải file exe tại: {download_url}")
        return

    print(f"\n⏳ Đang tải bản cập nhật v{new_version}...")
    new_exe_path = CURRENT_EXE + ".new"
    
    def reporthook(blocknum, blocksize, totalsize):
        read = blocknum * blocksize
        if totalsize > 0:
            percent = min(100, int(read * 100 / totalsize))
            mb_read = read / (1024 * 1024)
            mb_total = totalsize / (1024 * 1024)
            sys.stdout.write(f"\r📥 Đang tải: {percent}% ({mb_read:.1f}MB / {mb_total:.1f}MB)")
            sys.stdout.flush()

    try:
        urllib.request.urlretrieve(download_url, new_exe_path, reporthook)
        print("\n✅ Tải về hoàn tất! Đang khởi động lại ứng dụng...")
        
        bat_script = f"""@echo off
timeout /t 2 /nobreak > nul
move /y "{new_exe_path}" "{CURRENT_EXE}"
start "" "{CURRENT_EXE}"
del "%~f0"
"""
        updater_bat = os.path.join(APP_DIR, "updater.bat")
        with open(updater_bat, "w", encoding="utf-8") as f:
            f.write(bat_script)
            
        subprocess.Popen(["cmd.exe", "/c", updater_bat], close_fds=True)
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Lỗi khi tự động cập nhật: {e}")
        print(f"👉 Bạn có thể tải thủ công tại: {download_url}")
        if os.path.exists(new_exe_path):
            try: os.remove(new_exe_path)
            except: pass

def check_for_updates():
    """Tự động kiểm tra bản cập nhật mới nhất từ GitHub Releases"""
    try:
        print("🔍 Đang kiểm tra bản cập nhật mới...", flush=True)
        url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        req = urllib.request.Request(url, headers={"User-Agent": "ShiftLink-App"})
        with urllib.request.urlopen(req, timeout=4) as response:
            if response.status == 200:
                data = json.loads(response.read().decode())
                latest_tag = data.get("tag_name", "").lstrip("v")
                body = data.get("body", "Không có ghi chú cập nhật.")
                
                download_url = None
                for asset in data.get("assets", []):
                    if asset.get("name", "").endswith(".exe"):
                        download_url = asset.get("browser_download_url")
                        break

                if latest_tag and is_newer_version(latest_tag, CURRENT_VERSION) and download_url:
                    print("\n" + "=" * 65)
                    print(f"🎉 ĐÃ CÓ PHIÊN BẢN MỚI: v{latest_tag} (Phiên bản hiện tại: v{CURRENT_VERSION})")
                    print(f"📝 Nội dung mới:\n{body}")
                    print("=" * 65)
                    
                    choice = input("👉 Bạn có muốn tự động cập nhật ngay bây giờ không? (y/n, mặc định y): ").strip().lower()
                    if choice in ['', 'y', 'yes']:
                        perform_update(download_url, latest_tag)
                else:
                    print(f"✅ Bạn đang sử dụng phiên bản mới nhất (v{CURRENT_VERSION}).\n", flush=True)
    except Exception:
        pass

def wait_for_user_login(page):
    """Kiểm tra và chờ người dùng đăng nhập (không giới hạn thời gian)"""
    time.sleep(1.5)
    
    user_profile = page.locator("#user-profile")
    user_display_name = page.locator("#user-display-name")
    profile_classes = user_profile.get_attribute("class") or ""
    
    if user_profile.is_visible() and "hidden" not in profile_classes:
        username = user_display_name.inner_text().strip() or "User"
        print(f"👤 Tài khoản đã đăng nhập: {username}")
        return True

    print("\n" + "-" * 65)
    print("⚠️ CHƯA ĐĂNG NHẬP!")
    print("👉 Xin mời bạn đăng ký / đăng nhập trực tiếp trên cửa sổ Chrome vừa mở...")
    print("⏳ Hệ thống đang chờ bạn đăng nhập (Không giới hạn thời gian)...")
    print("-" * 65)

    waited_seconds = 0
    while True:
        time.sleep(1)
        waited_seconds += 1
        
        try:
            if page.is_closed():
                print("\n❌ Bạn đã đóng trình duyệt Chrome. Dừng chương trình.")
                return False
                
            p_classes = user_profile.get_attribute("class") or ""
            if user_profile.is_visible() and "hidden" not in p_classes:
                username = user_display_name.inner_text().strip() or "User"
                print(f"\n🎉 ĐĂNG NHẬP THÀNH CÔNG! Chào mừng {username}.")
                return True
        except Exception:
            pass

        if waited_seconds % 2 == 0:
            mins, secs = divmod(waited_seconds, 60)
            time_str = f"{mins:02d}:{secs:02d}" if mins > 0 else f"{secs}s"
            sys.stdout.write(f"\r⏳ Đang đợi bạn đăng nhập trên Chrome... ({time_str})")
            sys.stdout.flush()

def extract_links_from_sheet(sheet):
    """Trích xuất danh sách link và xác định cột link trong 1 sheet"""
    best_col = None
    max_links = 0
    for col in range(1, sheet.max_column + 1):
        link_count = 0
        for row in range(2, min(50, sheet.max_row + 1)):
            val = str(sheet.cell(row=row, column=col).value or '')
            if re.search(r'https?://[^\s]+', val):
                link_count += 1
        if link_count > max_links:
            max_links = link_count
            best_col = col

    if not best_col:
        return None, []

    links = []
    for r in range(2, sheet.max_row + 1):
        cell_val = str(sheet.cell(row=r, column=best_col).value or '')
        match = re.search(r'https?://[^\s]+', cell_val)
        if match:
            links.append((r, match.group(0), cell_val))
            
    return best_col, links

def generate_short_slug(original_url, attempt=0):
    slug_part = re.sub(r'^https?://[^/]+(?:/[^/]+)*/', '', original_url.rstrip('/'))
    target_len = random.randint(15, 21)
    
    collected_chars = []
    char_count = 0
    for ch in reversed(slug_part):
        collected_chars.append(ch)
        if ch != '-':
            char_count += 1
        if char_count >= target_len:
            break
            
    result = ''.join(reversed(collected_chars))
    result = re.sub(r'^[^a-zA-Z0-9]+', '', result)
    result = re.sub(r'[^a-zA-Z0-9]+$', '', result)
    
    if attempt > 1:
        suffix = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=2))
        result = f"{result[:18]}-{suffix}"
        
    return result

def run_automation(excel_path, selected_domain="nextpart2.online", max_items=None, target_sheet_names=None):
    excel_path = excel_path.strip('\'"')
    if not os.path.exists(excel_path):
        print(f"❌ Lỗi: Không tìm thấy file Excel tại '{excel_path}'", flush=True)
        return False

    wb = openpyxl.load_workbook(excel_path)
    
    sheets_to_process = []
    if target_sheet_names:
        sheets_to_process = [s for s in target_sheet_names if s in wb.sheetnames]
    else:
        sheets_to_process = wb.sheetnames

    tasks = []
    total_links_all = 0

    for sname in sheets_to_process:
        sheet = wb[sname]
        col_idx, links = extract_links_from_sheet(sheet)
        if links:
            out_col = None
            for c in range(1, sheet.max_column + 1):
                if str(sheet.cell(1, c).value or '').strip().lower() == "link da rut gon":
                    out_col = c
                    break
            if not out_col:
                out_col = sheet.max_column + 1
                sheet.cell(row=1, column=out_col, value="Link da rut gon")
                
            items = links[:max_items] if max_items else links
            tasks.append((sname, sheet, col_idx, out_col, items))
            total_links_all += len(items)

    if not tasks:
        print("❌ Không tìm thấy đường link nào trong các sheet đã chọn!", flush=True)
        return False

    print(f"\n📊 Bắt đầu xử lý: {len(tasks)} Sheet với tổng cộng {total_links_all} dòng dữ liệu...")
    print("=" * 65, flush=True)

    url_cache = {}

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            channel="chrome",
            headless=False,
            args=["--start-maximized"],
            no_viewport=True
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto(TARGET_URL, wait_until="domcontentloaded")

        if not wait_for_user_login(page):
            context.close()
            return False

        processed_count = 0
        for sname, sheet, col_idx, out_col, items in tasks:
            print(f"\n--- 📂 ĐANG XỬ LÝ SHEET: '{sname}' ({len(items)} links) ---", flush=True)
            for row_idx, original_url, full_cell_text in items:
                processed_count += 1
                
                # TRƯỜNG HỢP 1: LINK GỐC ĐÃ TỪNG RÚT GỌN TRƯỚC ĐÓ (DÙNG LẠI CACHE)
                if original_url in url_cache:
                    cached_val = url_cache[original_url]
                    sheet.cell(row=row_idx, column=out_col, value=cached_val)
                    print(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ⚡ [DÙNG LẠI] {cached_val}", flush=True)
                    continue

                # TRƯỜNG HỢP 2: RÚT GỌN MỚI TRÊN WEB VÀ THEO DÕI THỜI GIAN CHỜ BÁO THÀNH CÔNG
                success = False
                for attempt in range(5):
                    slug = generate_short_slug(original_url, attempt=attempt)
                    
                    # 1. Điền Form trên giao diện
                    page.locator("#original-url").fill(original_url)
                    page.locator("#domain-select").select_option(label=selected_domain)
                    page.locator("#short-path").fill(slug)

                    # Dừng nhẹ 0.4s để mắt nhìn thấy dữ liệu vừa điền
                    time.sleep(0.4)

                    try:
                        # 2. Nhấn nút "Rút Gọn Link"
                        submit_btn = page.locator("#shorten-form button[type='submit']")
                        
                        with page.expect_response(lambda res: "/api/links" in res.url and res.request.method == "POST", timeout=12000) as resp_info:
                            submit_btn.click()

                        response = resp_info.value
                        
                        # 3. Chờ thông báo phản hồi (Toast notification trên giao diện web)
                        toast_msg = "Thành công"
                        try:
                            toast_elem = page.locator("#toast.show")
                            toast_elem.wait_for(state="visible", timeout=3000)
                            toast_msg = toast_elem.inner_text().strip()
                        except Exception:
                            pass

                        if response.status in [200, 201]:
                            res_data = response.json()
                            created_path = res_data.get("shortPath", slug)
                            shortened_url = f"https://{selected_domain}/{created_path}"
                            final_text = f"watch full here 👉: {shortened_url}"
                            
                            url_cache[original_url] = final_text
                            sheet.cell(row=row_idx, column=out_col, value=final_text)
                            
                            # 4. In thông báo thành công sau khi trang web đã phản hồi
                            print(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> 📢 [{toast_msg}] ✅ {final_text}", flush=True)
                            success = True
                            
                            # 5. Dừng nghỉ tự nhiên 1.5s giữa các link để web kịp cập nhật bảng và người dùng dễ quan sát
                            time.sleep(1.5)
                            break
                        else:
                            print(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ⚠️ Web báo: {toast_msg}, đang thử lại slug khác...", flush=True)
                            time.sleep(1.0)
                    except Exception as e:
                        time.sleep(1.0)

                if not success:
                    print(f"[{processed_count:03d}/{total_links_all:03d}] [{sname}] Dòng #{row_idx:<3d} -> ❌ Thất bại sau 5 lần thử.", flush=True)

        context.close()

    if tasks:
        wb.active = tasks[0][1]

    dir_name = os.path.dirname(excel_path)
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    out_file = os.path.join(dir_name, f"{base_name}_rutgon.xlsx")
    wb.save(out_file)

    print("=" * 65, flush=True)
    print("🎉 HOÀN TẤT TẤT CẢ CÁC SHEET!", flush=True)
    print(f"💾 File kết quả đã lưu tại: {out_file}", flush=True)
    return True

def main():
    print("*" * 65)
    print(f"   CÔNG CỤ TỰ ĐỘNG HÓA RÚT GỌN LINK SHIFTLINK (v{CURRENT_VERSION})   ")
    print("*" * 65)
    
    check_for_updates()

    excel_path = input("👉 Kéo thả file Excel vào đây (hoặc dán đường dẫn): ").strip('\'"')
    if not excel_path:
        default_file = r"C:\Users\TP\Downloads\đợt 2.xlsx"
        if os.path.exists(default_file):
            print(f"Sử dụng file mặc định: {default_file}")
            excel_path = default_file
        else:
            print("❌ Vui lòng nhập đường dẫn file Excel hợp lệ!")
            input("\nNhấn Enter để thoát...")
            return

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"❌ Lỗi khi đọc file Excel: {e}")
        input("\nNhấn Enter để thoát...")
        return

    target_sheets = None
    if len(sheet_names) > 1:
        print("\n" + "=" * 65)
        print(f"📄 File Excel này có {len(sheet_names)} Sheet:")
        print("   [0] 🌟 RÚT GỌN TẤT CẢ CÁC SHEET")
        for idx, sname in enumerate(sheet_names, start=1):
            print(f"   [{idx}] Sheet '{sname}'")
        print("=" * 65)
        
        sheet_choice = input(f"👉 Chọn Sheet muốn rút gọn (Nhập 0 để rút gọn tất cả, hoặc 1, 2,... {len(sheet_names)}; mặc định 0): ").strip()
        if sheet_choice.isdigit():
            s_idx = int(sheet_choice)
            if 1 <= s_idx <= len(sheet_names):
                target_sheets = [sheet_names[s_idx - 1]]
                print(f"✅ Đã chọn rút gọn riêng Sheet: '{target_sheets[0]}'")
            else:
                print("✅ Đã chọn rút gọn TẤT CẢ các Sheet.")
        else:
            print("✅ Đã chọn rút gọn TẤT CẢ các Sheet.")
    else:
        print(f"📄 File có 1 Sheet: '{sheet_names[0]}'")

    domain = input("\n👉 Chọn tên miền (Nhấn Enter để dùng mặc định: nextpart2.online): ").strip()
    if not domain:
        domain = "nextpart2.online"

    limit_str = input("👉 Số lượng dòng muốn xử lý (Nhấn Enter để xử lý TOÀN BỘ): ").strip()
    limit = int(limit_str) if limit_str.isdigit() else None

    run_automation(excel_path, selected_domain=domain, max_items=limit, target_sheet_names=target_sheets)

    print("\n" + "=" * 65)
    input("👉 Nhấn phím Enter để kết thúc...")

if __name__ == "__main__":
    main()
